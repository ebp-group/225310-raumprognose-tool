"""Raumprognose Tool – Flet Desktop Application.

Run with:
    flet run app/flet_app.py
    Or directly:
    python app/flet_app.py

Assets (splash screen, icon) are served from the ``app/assets/`` directory.
"""

from __future__ import annotations

import base64
import io
import os
import sys
import asyncio
import re
import zipfile
from numbers import Real
from pathlib import Path
from typing import Any
import logging
import flet as ft
import flet_datatable2 as fdt
import matplotlib
import yaml

matplotlib.use("Agg")  # noqa: E402 – must be set before importing pyplot
import matplotlib.pyplot as plt  # noqa: E402
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Ensure the app directory is on the path for sibling imports
sys.path.insert(0, os.path.dirname(__file__))

from calculations import (
    area_by_eigentumsform,
    current_area_by_nutzungsart,
    future_demand,
    surplus_deficit,
)
from data_loader import load_gebaeude_raeume, load_nutzungsfaktoren, load_studierende


def get_assets_dir() -> Path:
    default_assets_dir = Path(__file__).parent / "assets"  # fallback for local runs
    return Path(os.environ.get("FLET_ASSETS_DIR", str(default_assets_dir))).resolve()


def _resolve_app_version() -> str:
    # 1. Bundled version file written by the CI build step
    _version_file = get_assets_dir() / "version.txt"
    if _version_file.is_file():
        _v = _version_file.read_text(encoding="utf-8").strip()
        if _v:
            return _v
    # 2. Installed-package metadata (works in normal pip/uv environments)
    try:
        from importlib.metadata import version as _pkg_version

        return _pkg_version("raumprognose-tool")
    except Exception:
        pass
    # 3. Hardcoded fallback
    return "0.1.0"


APP_VERSION = _resolve_app_version()

APP_NAME = "Raumprognose Tool"
APP_COPYRIGHT_YEAR = "2026"
APP_COPYRIGHT = f"© {APP_COPYRIGHT_YEAR} EBP"
APP_LICENSE_NAME = "BSD 3-Clause License"
APP_LICENSE_TEXT = (
    "Redistribution and use in source and binary forms, with or without "
    "modification, are permitted provided that the following conditions are met:\n\n"
    "1. Redistributions of source code must retain the above copyright notice, "
    "this list of conditions and the following disclaimer.\n\n"
    "2. Redistributions in binary form must reproduce the above copyright notice, "
    "this list of conditions and the following disclaimer in the documentation "
    "and/or other materials provided with the distribution.\n\n"
    "3. Neither the name of the copyright holder nor the names of its contributors "
    "may be used to endorse or promote products derived from this software without "
    "specific prior written permission.\n\n"
    'THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS" '
    "AND WITHOUT WARRANTY OF ANY KIND."
)
APP_DOCUMENTATION_URL = "https://ebp-group.github.io/225310-raumprognose-tool/"

SPLASH_DURATION_SECONDS = 2
SPLASH_FADE_OUT_MS = 300

log = logging.getLogger(__name__)
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s"
)
log.setLevel(logging.DEBUG)


def _load_nutzungsart_config() -> tuple[dict[str, str], dict[str, str]]:
    """Load display labels and colors for Nutzungsarten from config.yml."""
    config_path = get_assets_dir() / "config.yml"
    try:
        with config_path.open("r", encoding="utf-8") as f:
            data = yaml.safe_load(f) or {}
    except FileNotFoundError:
        log.warning("Nutzungsart config not found: %s", config_path)
        return {}, {}
    except Exception as exc:
        log.warning("Failed to load Nutzungsart config from %s: %s", config_path, exc)
        return {}, {}

    display_map: dict[str, str] = {}
    color_map: dict[str, str] = {}
    for entry in data.get("nutzungsarten", []):
        raw_name = entry.get("name")
        display_name = entry.get("column_name")
        color = entry.get("color")
        if not raw_name:
            continue
        if display_name:
            display_map[str(raw_name)] = str(display_name)
        if color:
            normalized_color = str(color).lstrip("#").upper()
            if re.fullmatch(r"[0-9A-F]{6}", normalized_color):
                color_map[str(raw_name)] = normalized_color
            else:
                log.warning(
                    "Invalid color '%s' for Nutzungsart '%s' in %s",
                    color,
                    raw_name,
                    config_path,
                )

    return display_map, color_map


NUTZUNGSART_DISPLAY_MAP, NUTZUNGSART_COLOR_MAP = _load_nutzungsart_config()

# ── UI helper functions ───────────────────────────────────────────────────────


def _replace_thousands_commas(text: str) -> str:
    """Replace thousands-separator commas with apostrophes in numeric strings."""
    return re.sub(
        r"-?\d{1,3}(?:,\d{3})+(?:\.\d+)?",
        lambda match: match.group(0).replace(",", "'"),
        text,
    )


def _format_display_value(value: Any, round_to: int|float = 1) -> str:
    """Format values for UI display with apostrophe thousands separators."""
    if pd.isna(value):
        return ""
    if isinstance(value, Real) and not isinstance(value, bool) and isinstance(round_to, int):
        value = round(value / round_to) * round_to
        return f"{value:,}".replace(",", "'")
    if isinstance(value, Real) and not isinstance(value, bool):
        value = round(value / round_to) * round_to
        return f"{value:,.1f}".replace(",", "'")
    return _replace_thousands_commas(str(value))


def _map_nutzungsart_values(series: pd.Series) -> pd.Series:
    """Map technical Nutzungsart values to configured display labels."""
    return series.map(NUTZUNGSART_DISPLAY_MAP).fillna(series)


def _apply_nutzungsart_labels(df: pd.DataFrame) -> pd.DataFrame:
    """Return a copy with mapped Nutzungsart labels when that column exists."""
    if "Nutzungsart" not in df.columns:
        return df.copy()
    mapped = df.copy()
    mapped["Nutzungsart"] = _map_nutzungsart_values(mapped["Nutzungsart"])
    return mapped


def _df_to_datatable(df: pd.DataFrame, max_rows: int = 200, round_to: int|float = 1) -> ft.DataTable:
    """Convert a pandas DataFrame to a Flet DataTable widget."""
    columns = [
        ft.DataColumn(label=ft.Text(str(col), weight=ft.FontWeight.BOLD))
        for col in df.columns
    ]
    rows = []
    for _, row in df.head(max_rows).iterrows():
        cells = [ft.DataCell(ft.Text(_format_display_value(val, round_to=round_to))) for val in row]
        rows.append(fdt.DataRow2(cells=cells))
    return ft.DataTable(
        columns=columns,
        rows=rows,
        border=ft.Border.all(1, ft.Colors.GREY_300),
        heading_row_color=ft.Colors.BLUE_GREY_50,
        horizontal_lines=ft.BorderSide(1, ft.Colors.GREY_200),
    )


def _metric_card(label: str, value: Any, is_surplus: bool) -> ft.Card:
    """Build a compact metric display card."""
    formatted_value = _format_display_value(value)
    return ft.Card(
        content=ft.Container(
            content=ft.Column(
                [
                    ft.Text(label, size=14, color=ft.Colors.GREY_600),
                    ft.Text(formatted_value, size=20, weight=ft.FontWeight.BOLD),
                    ft.Text(
                        "Überschuss" if is_surplus else "Defizit",
                        size=12,
                        color=ft.Colors.GREEN if is_surplus else ft.Colors.RED,
                    ),
                ],
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                spacing=4,
            ),
            padding=16,
            alignment=ft.Alignment.CENTER,
        ),
        elevation=2,
    )


def _fig_to_base64(fig: plt.Figure) -> str:
    """Render a matplotlib figure to a base64-encoded PNG string."""
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    buf.seek(0)
    return base64.b64encode(buf.getvalue()).decode("utf-8")


def _build_png_zip(figs: list[tuple[str, plt.Figure]]) -> bytes:
    """Build a ZIP archive containing PNG renders of multiple figures."""
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for filename, fig in figs:
            png_buf = io.BytesIO()
            fig.savefig(png_buf, format="png", dpi=150, bbox_inches="tight")
            archive.writestr(filename, png_buf.getvalue())
    return zip_buf.getvalue()


# ── Chart builders ────────────────────────────────────────────────────────────


def _create_students_chart(df_studierende: pd.DataFrame) -> plt.Figure:
    """Line chart with one line per category over time."""
    fig, ax = plt.subplots(figsize=(8, 4))

    required_columns = {"Jahr", "Kategorie", "Anzahl"}
    missing_columns = sorted(required_columns - set(df_studierende.columns))
    if missing_columns:
        raise ValueError(
            "Studierenden-Daten müssen im Long-Format vorliegen "
            f"(fehlende Spalten: {missing_columns})."
        )

    chart_df = (
        df_studierende.pivot_table(
            index="Jahr",
            columns="Kategorie",
            values="Anzahl",
            aggfunc="sum",
        )
        .reset_index()
        .sort_values("Jahr")
    )

    categories = ["Studierende", "Forschung", "Services", "Stundenlohn"]
    palette = plt.get_cmap("tab10")

    for idx, label in enumerate(categories):
        if label not in chart_df.columns:
            continue
        ax.plot(
            chart_df["Jahr"],
            chart_df[label],
            marker="o",
            linewidth=2,
            label=label,
            color=palette(idx % palette.N),
        )

    ax.set_xlabel("Jahr")
    ax.set_ylabel("Anzahl")
    ax.set_title("Entwicklung nach Kategorie")
    if ax.lines:
        ax.legend()
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    return fig


def _create_demand_chart(df_demand: pd.DataFrame, scenario: str) -> plt.Figure:
    """Grouped bar chart of area demand by usage type and year."""
    fig, ax = plt.subplots(figsize=(10, 5))
    df_demand = _apply_nutzungsart_labels(df_demand)
    # years = sorted(df_demand["jahr"].unique())
    years = [2026, 2030, 2040, 2050]

    nutzungsarten = sorted(
        df_demand[df_demand["Bedarf_m2"] > 0]["Nutzungsart"].unique()
    )
    x = range(len(nutzungsarten))
    width = 0.8 / max(len(years), 1)

    for i, year in enumerate(years):
        df_year = df_demand[df_demand["Jahr"] == year]
        values = []
        for n in nutzungsarten:
            subset = df_year[df_year["Nutzungsart"] == n]["Bedarf_m2"]
            values.append(subset.values[0] if len(subset) > 0 else 0)
        offset = (i - len(years) / 2 + 0.5) * width
        ax.bar([xi + offset for xi in x], values, width, label=str(year))

    ax.set_xlabel("Nutzungsart")
    ax.set_ylabel("Bedarf (m²)")
    ax.set_title(f"Flächenbedarf – Szenario: {scenario}")
    ax.set_xticks(list(x))
    ax.set_xticklabels(nutzungsarten, rotation=45, ha="right")
    ax.legend()
    ax.grid(True, alpha=0.3, axis="y")
    fig.tight_layout()
    return fig


def _create_surplus_deficit_charts(df_sd: pd.DataFrame) -> list[tuple[int, plt.Figure]]:
    """One bar chart per forecast year showing surplus/deficit by usage type."""
    df_sd = _apply_nutzungsart_labels(df_sd)
    # years = sorted(df_sd["jahr"].unique())
    years = [2026, 2030, 2040, 2050]
    figs: list[tuple[int, plt.Figure]] = []
    for year in years:
        fig, ax = plt.subplots(figsize=(5, 4))
        df_year = df_sd[df_sd["Jahr"] == year]
        df_year = df_year.dropna(subset=["Differenz_m2"])
        colors = ["#2ecc71" if v >= 0 else "#e74c3c" for v in df_year["Differenz_m2"]]
        ax.bar(df_year["Nutzungsart"], df_year["Differenz_m2"], color=colors)
        ax.set_xlabel("Nutzungsart")
        ax.set_ylabel("Differenz (m²)")
        ax.set_title(str(year))
        ax.axhline(y=0, color="black", linewidth=0.5)
        ax.tick_params(axis="x", rotation=45)
        ax.grid(True, alpha=0.3, axis="y")
        fig.tight_layout()
        figs.append((year, fig))
    return figs


def _create_eigentumsform_chart(df_gebaeude: pd.DataFrame) -> plt.Figure:
    """Stacked bar chart of area by ownership type (Eigentumsform) over selected years."""
    years = [2025, 2026, 2030, 2040]
    df = area_by_eigentumsform(df_gebaeude, years)

    # Pivot: index = Eigentumsform, columns = Jahr
    pivot = df.pivot_table(
        index="Eigentumsform",
        columns="Jahr",
        values="Fläche",
        aggfunc="sum",
        fill_value=0,
    )
    pivot = pivot.reindex(columns=years, fill_value=0)

    eigentumsformen = pivot.index.tolist()
    palette = plt.get_cmap("tab20")
    x = range(len(years))

    fig, ax = plt.subplots(figsize=(9, 5))
    bottom = [0.0] * len(years)
    for idx, eigentumsform in enumerate(eigentumsformen):
        values = [float(pivot.loc[eigentumsform, year]) for year in years]
        ax.bar(
            x,
            values,
            bottom=bottom,
            label=str(eigentumsform),
            color=palette(idx % palette.N),
        )
        bottom = [b + v for b, v in zip(bottom, values)]

    ax.set_xlabel("Jahr")
    ax.set_ylabel("Fläche (m²)")
    ax.set_title("Fläche nach Eigentumsform")
    ax.set_xticks(list(x))
    ax.set_xticklabels([str(y) for y in years])
    if eigentumsformen:
        ax.legend(title="Eigentumsform", bbox_to_anchor=(1.01, 1), loc="upper left")
    ax.grid(True, alpha=0.3, axis="y")
    fig.tight_layout()
    return fig


# ── Excel export builder ─────────────────────────────────────────────────────


def _build_excel(
    df_results: pd.DataFrame,
    df_stud: pd.DataFrame,
    df_dem: pd.DataFrame,
) -> bytes:
    """Build a styled multi-sheet Excel workbook and return it as bytes.

    Args:
        df_results: Full surplus/deficit DataFrame.
        df_stud: Student numbers DataFrame.
        df_dem: Future demand DataFrame.

    Returns:
        Raw bytes of the ``.xlsx`` workbook.
    """
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    center = Alignment(horizontal="center")

    def _write_sheet(
        ws: Any,
        df: pd.DataFrame,
        title: str,
        diff_col: str | None = None,
        nutzungsart_colors: pd.Series | None = None,
    ) -> None:
        ws.title = title
        header_row_index = 1
        first_data_row_index = header_row_index + 1
        nutzungsart_col_idx = (
            list(df.columns).index("Nutzungsart") + 1
            if nutzungsart_colors is not None and "Nutzungsart" in df.columns
            else None
        )
        for r_idx, row in enumerate(
            dataframe_to_rows(df, index=False, header=True), start=1
        ):
            ws.append(row)
            if r_idx == header_row_index:
                for cell in ws[r_idx]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center
            elif diff_col and r_idx >= first_data_row_index:
                col_names = list(df.columns)
                if diff_col in col_names:
                    c_idx = col_names.index(diff_col) + 1
                    val = ws.cell(row=r_idx, column=c_idx).value
                    try:
                        if float(val) >= 0:
                            ws.cell(row=r_idx, column=c_idx).fill = green_fill
                        else:
                            ws.cell(row=r_idx, column=c_idx).fill = red_fill
                    except (TypeError, ValueError):
                        pass

            if r_idx >= first_data_row_index and nutzungsart_col_idx is not None:
                row_color = nutzungsart_colors.iloc[r_idx - first_data_row_index]
                if pd.notna(row_color):
                    ws.cell(row=r_idx, column=nutzungsart_col_idx).fill = PatternFill(
                        "solid",
                        fgColor=str(row_color),
                    )

    df_results_export = df_results.copy()
    df_results_nutzungsart_colors = (
        df_results_export["Nutzungsart"].map(NUTZUNGSART_COLOR_MAP)
        if "Nutzungsart" in df_results_export.columns
        else None
    )
    df_results_export = _apply_nutzungsart_labels(df_results_export)

    df_dem_export = df_dem.copy()
    df_dem_nutzungsart_colors = (
        df_dem_export["Nutzungsart"].map(NUTZUNGSART_COLOR_MAP)
        if "Nutzungsart" in df_dem_export.columns
        else None
    )
    df_dem_export = _apply_nutzungsart_labels(df_dem_export)

    ws1 = wb.active
    _write_sheet(
        ws1,
        df_results_export,
        "Ergebnisse",
        diff_col="Differenz_m2",
        nutzungsart_colors=df_results_nutzungsart_colors,
    )

    ws2 = wb.create_sheet()
    _write_sheet(ws2, df_stud, "Studierende")

    ws3 = wb.create_sheet()
    _write_sheet(
        ws3,
        df_dem_export,
        "Flächenbedarf",
        nutzungsart_colors=df_dem_nutzungsart_colors,
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_excel_rounded(df_sd: pd.DataFrame) -> bytes:
    """Build a wide-format Excel workbook with areas rounded to the nearest 5.

    The sheet has one row per usage type (``Nutzungsart``) and three columns
    per forecast year: *IST* (available area), *SOLL* (required area) and
    *Differenz*.  All values are rounded to the nearest 5 and contain no
    decimal places.

    Args:
        df_sd: Output of :func:`surplus_deficit` with columns ``Nutzungsart``,
            ``Jahr``, ``Fläche``, ``Bedarf_m2``, ``Differenz_m2``.

    Returns:
        Raw bytes of the ``.xlsx`` workbook.
    """

    def _round5(x: float) -> int:
        """Round *x* to the nearest multiple of 5."""
        return int(round(x / 5) * 5)

    # Round numeric columns
    df = df_sd.copy()
    for col in ("Fläche", "Bedarf_m2", "Differenz_m2"):
        df[col] = df[col].apply(lambda v: _round5(v) if pd.notna(v) else None)

    years = sorted(df["Jahr"].unique())

    # Build wide DataFrame: index = Nutzungsart, columns = (year, metric)
    usage_types = df["Nutzungsart"].unique()
    col_tuples = [(yr, label) for yr in years for label in ("IST", "SOLL", "Differenz")]
    wide = pd.DataFrame(index=usage_types, columns=pd.MultiIndex.from_tuples(col_tuples))
    wide.index.name = "Nutzungsart"

    for _, row in df.iterrows():
        yr = row["Jahr"]
        nt = row["Nutzungsart"]
        wide.loc[nt, (yr, "IST")] = row["Fläche"]
        wide.loc[nt, (yr, "SOLL")] = row["Bedarf_m2"]
        wide.loc[nt, (yr, "Differenz")] = row["Differenz_m2"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Gerundete Flächen"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    subheader_fill = PatternFill("solid", fgColor="2E75B6")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    center = Alignment(horizontal="center")

    # Row 1: "Nutzungsart" header + year group headers (merged over 3 cols each)
    ws.cell(row=1, column=1, value="Nutzungsart")
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = center

    for i, yr in enumerate(years):
        col_start = 2 + i * 3
        ws.cell(row=1, column=col_start, value=int(yr))
        ws.cell(row=1, column=col_start).font = header_font
        ws.cell(row=1, column=col_start).fill = header_fill
        ws.cell(row=1, column=col_start).alignment = center
        ws.merge_cells(
            start_row=1, start_column=col_start,
            end_row=1, end_column=col_start + 2,
        )

    # Row 2: sub-headers IST / SOLL / Differenz per year
    ws.cell(row=2, column=1, value="Nutzungsart")
    ws.cell(row=2, column=1).font = header_font
    ws.cell(row=2, column=1).fill = header_fill
    ws.cell(row=2, column=1).alignment = center

    for i, _yr in enumerate(years):
        for j, label in enumerate(("IST", "SOLL", "Differenz")):
            col = 2 + i * 3 + j
            cell = ws.cell(row=2, column=col, value=label)
            cell.font = header_font
            cell.fill = subheader_fill
            cell.alignment = center

    # Data rows
    for r_idx, (nt, row_data) in enumerate(wide.iterrows(), start=3):
        raw_name = str(nt)
        display_name = NUTZUNGSART_DISPLAY_MAP.get(raw_name, raw_name)
        nt_cell = ws.cell(row=r_idx, column=1, value=display_name)
        nt_color = NUTZUNGSART_COLOR_MAP.get(raw_name)
        if nt_color:
            nt_cell.fill = PatternFill("solid", fgColor=nt_color)
        for i, yr in enumerate(years):
            for j, metric in enumerate(("IST", "SOLL", "Differenz")):
                col = 2 + i * 3 + j
                val = row_data[(yr, metric)]
                cell = ws.cell(row=r_idx, column=col, value=val)
                if metric == "Differenz" and val is not None:
                    try:
                        cell.fill = green_fill if float(val) >= 0 else red_fill
                    except (TypeError, ValueError):
                        pass

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── About dialog ─────────────────────────────────────────────────────────────


def _show_about_dialog(page: ft.Page) -> None:
    """Open a modal dialog with application information."""

    dlg = ft.AlertDialog(
        modal=True,
        title=ft.Text("Über diese Applikation", weight=ft.FontWeight.BOLD),
        content=ft.Container(
            content=ft.Column(
                [
                    ft.Image(
                        src="icon.png",
                        width=80,
                        height=80,
                        fit=ft.BoxFit.CONTAIN,
                    ),
                    ft.Text(
                        APP_NAME,
                        size=18,
                        weight=ft.FontWeight.BOLD,
                    ),
                    ft.Text(
                        f"Version {APP_VERSION}",
                        size=13,
                        color=ft.Colors.GREY_600,
                    ),
                    ft.Text(
                        APP_COPYRIGHT,
                        size=13,
                    ),
                    ft.Divider(),
                    ft.Text(
                        APP_LICENSE_NAME,
                        weight=ft.FontWeight.BOLD,
                        size=13,
                    ),
                    ft.Container(
                        content=ft.Text(
                            APP_LICENSE_TEXT,
                            size=11,
                            color=ft.Colors.GREY_700,
                        ),
                        padding=ft.Padding(8, 4, 8, 4),
                        bgcolor=ft.Colors.GREY_100,
                        border_radius=4,
                        width=440,
                    ),
                ],
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                spacing=8,
            ),
            padding=ft.Padding(8, 0, 8, 0),
        ),
        actions=[
            ft.TextButton(
                "Schliessen",
                on_click=lambda _: page.pop_dialog(),
            )
        ],
        actions_alignment=ft.MainAxisAlignment.END,
    )
    page.show_dialog(dlg)


# ── Main application ─────────────────────────────────────────────────────────


def main(page: ft.Page) -> None:
    """Build and display the Raumprognose desktop application."""

    page.title = "🏛️ Raumprognose Tool"
    page.window.width = 1400
    page.window.height = 900
    page.theme_mode = ft.ThemeMode.LIGHT
    page.padding = 0

    async def _open_documentation() -> None:
        await ft.UrlLauncher().launch_url(
            APP_DOCUMENTATION_URL,
            mode=ft.LaunchMode.IN_APP_BROWSER_VIEW,
            browser_configuration=ft.BrowserConfiguration(show_title=True),
        )

    page.appbar = ft.AppBar(
        title=ft.Text("🏛️ Raumprognose Tool"),
        bgcolor=ft.Colors.SURFACE,
        actions=[
            ft.MenuBar(
                controls=[
                    ft.SubmenuButton(
                        content=ft.Text("Hilfe"),
                        controls=[
                            ft.MenuItemButton(
                                content=ft.Text("Dokumentation"),
                                on_click=_open_documentation,
                            ),
                            ft.MenuItemButton(
                                content=ft.Text("Über"),
                                on_click=lambda _: _show_about_dialog(page),
                            ),
                        ],
                    ),
                ],
                style=ft.MenuStyle(bgcolor=ft.Colors.SURFACE),
            ),
        ],
    )

    # ── Splash screen overlay ─────────────────────────────────────────────

    splash_image = ft.Image(src="splash.png", fit=ft.BoxFit.CONTAIN)
    splash_overlay = ft.Container(
        content=splash_image,
        alignment=ft.Alignment.CENTER,
        bgcolor=ft.Colors.WHITE,
        expand=True,
        opacity=1.0,
        animate_opacity=ft.Animation(SPLASH_FADE_OUT_MS, ft.AnimationCurve.EASE_OUT),
    )
    page.overlay.append(splash_overlay)
    page.update()

    async def dismiss_splash() -> None:
        """Fade out and remove the splash screen after a short delay."""
        await asyncio.sleep(SPLASH_DURATION_SECONDS)
        splash_overlay.opacity = 0
        splash_overlay.update
        page.update()
        # Wait for the fade-out animation to finish, then remove the overlay.
        await asyncio.sleep(SPLASH_FADE_OUT_MS / 1000 + 0.1)

        if splash_overlay in page.overlay:
            page.overlay.remove(splash_overlay)
            page.update()

    page.run_task(dismiss_splash)

    # ── Mutable application state ─────────────────────────────────────────

    # TODO: remove hardcoded paths and use file pickers instead
    base_path = Path(
        r"C:\Users\ods\OneDrive - EBP\CH_P_225310 - PE_TPF_UniSG - General\40_BEARBEITUNG\04_Auswertung\02_Datenmodell"
    )

    state: dict[str, Any] = {
        "df_gebaeude": None,
        "df_studierende": None,
        "df_faktoren": None,
        "scenario": None,
        "custom_gebaeude": (
            base_path / "260402_UniSG_Rauminventar_rev_260414.xlsx"
        ),  # TODO: default paths for testing only, remove later
        "custom_studierende": (
            base_path / "prognose_studierende_und_ma.xlsx"
        ),  # TODO: default paths for testing only, remove later
        "custom_faktoren": (
            base_path / "nutzungsfaktoren.xlsx"
        ),  # TODO: default paths for testing only, remove later
    }

    # ── Data loading ──────────────────────────────────────────────────────

    def load_all_data(e) -> bool:
        """Load all three datasets. Returns *True* on success."""
        if any(
            state[f"custom_{key}"] is None
            for key in ("gebaeude", "studierende", "faktoren")
        ):
            page.show_dialog(
                ft.SnackBar(
                    content=ft.Text("Bitte zuerst alle Dateien auswählen."),
                    bgcolor=ft.Colors.RED_400,
                )
            )
            return False

        state["df_gebaeude"] = None
        state["df_studierende"] = None
        state["df_faktoren"] = None

        _update_scenario_options()
        rebuild_content()
        page.update()
        try:
            state["df_gebaeude"] = load_gebaeude_raeume(state["custom_gebaeude"])
            state["df_studierende"] = load_studierende(state["custom_studierende"])
            state["df_faktoren"] = load_nutzungsfaktoren(state["custom_faktoren"])

            _update_scenario_options()
            page.update()

            page.show_dialog(
                ft.SnackBar(
                    content=ft.Text("Daten erfolgreich geladen."),
                    bgcolor=ft.Colors.GREEN_400,
                )
            )
            return True
        except Exception as exc:
            log.exception("Error loading data")
            page.show_dialog(
                ft.SnackBar(
                    content=ft.Text(f"Fehler beim Laden der Daten: {exc}"),
                    bgcolor=ft.Colors.RED_400,
                )
            )
            return False

    def run_calculations(e) -> None:
        """Run the calculations and update the content."""
        if (
            state["df_gebaeude"] is None
            or state["df_studierende"] is None
            or state["df_faktoren"] is None
        ):
            page.show_dialog(
                ft.SnackBar(
                    content=ft.Text("Bitte zuerst alle Daten laden."),
                    bgcolor=ft.Colors.RED_400,
                )
            )
            return
        rebuild_content()
        page.update()

    def get_results() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        """Return (current_area, demand, surplus_deficit) DataFrames."""
        years = sorted(state["df_studierende"]["Jahr"].unique().tolist())
        df_current = current_area_by_nutzungsart(state["df_gebaeude"], years)
        df_demand = future_demand(
            state["df_studierende"], state["df_faktoren"], state["scenario"]
        )
        df_sd = surplus_deficit(df_current, df_demand)
        return df_current, df_demand, df_sd

    # ── File picker ──────────────────────────────────────────────────────

    gebaeude_label = ft.Text("Keine Datei gewählt", size=12, italic=True)
    studierende_label = ft.Text("Keine Datei gewählt", size=12, italic=True)
    faktoren_label = ft.Text("Keine Datei gewählt", size=12, italic=True)

    async def _pick_file(key: str, label: ft.Text):
        """Open a file-picker dialog for one of the three input datasets.

        The picked file path is stored in ``state["custom_{key}"]`` and
        the sidebar label is updated accordingly.

        Args:
            key: State key suffix – one of ``"gebaeude"``, ``"studierende"``,
                or ``"faktoren"``.
            label: The sidebar :class:`ft.Text` label to update with the
                chosen file name.
        """
        files = await ft.FilePicker().pick_files(
            allowed_extensions=["xlsx"],
            file_type=ft.FilePickerFileType.CUSTOM,
            dialog_title=f"{key.capitalize()} (.xlsx)",
        )
        if files:
            state[f"custom_{key}"] = files[0].path
            label.value = Path(files[0].path).name
        else:
            state[f"custom_{key}"] = None
            label.value = "Keine Datei gewählt"

        _update_scenario_options()
        rebuild_content()
        page.update()

    async def load_inventory_file(e) -> None:
        await _pick_file("gebaeude", gebaeude_label)

    async def load_person_file(e) -> None:
        await _pick_file("studierende", studierende_label)

    async def load_faktoren_file(e) -> None:
        await _pick_file("faktoren", faktoren_label)

    # Export handlers
    async def _save_excel(_e):
        _, df_demand, df_sd = get_results()
        excel_bytes = _build_excel(df_sd, state["df_studierende"], df_demand)
        path = await ft.FilePicker().save_file(
            file_name=f"raumprognose_{state['scenario']}.xlsx",
            allowed_extensions=["xlsx"],
            file_type=ft.FilePickerFileType.CUSTOM,
        )
        if path:
            with open(path, "wb") as f:
                f.write(excel_bytes)
            page.show_dialog(ft.SnackBar(content=ft.Text(f"Gespeichert: {path}")))
            page.update()

    async def _save_excel_rounded(_e):
        _, _, df_sd = get_results()
        excel_bytes = _build_excel_rounded(df_sd)
        path = await ft.FilePicker().save_file(
            file_name=f"raumprognose_gerundet_{state['scenario']}.xlsx",
            allowed_extensions=["xlsx"],
            file_type=ft.FilePickerFileType.CUSTOM,
        )
        if path:
            with open(path, "wb") as f:
                f.write(excel_bytes)
            page.show_dialog(ft.SnackBar(content=ft.Text(f"Gespeichert: {path}")))
            page.update()

    async def _save_students_png(_e):
        fig = _create_students_chart(state["df_studierende"])
        path = await ft.FilePicker().save_file(
            file_name="studierende.png",
            allowed_extensions=["png"],
            file_type=ft.FilePickerFileType.CUSTOM,
        )
        if path:
            fig.savefig(path, format="png", dpi=150, bbox_inches="tight")
            page.show_dialog(ft.SnackBar(content=ft.Text(f"Gespeichert: {path}")))
            page.update()
        plt.close(fig)

    async def _save_demand_png(_e):
        _, df_demand, _ = get_results()
        fig = _create_demand_chart(df_demand, state["scenario"])
        path = await ft.FilePicker().save_file(
            file_name=f"flaechenbedarf_{state['scenario']}.png",
            allowed_extensions=["png"],
            file_type=ft.FilePickerFileType.CUSTOM,
        )
        if path:
            fig.savefig(path, format="png", dpi=150, bbox_inches="tight")
            page.show_dialog(ft.SnackBar(content=ft.Text(f"Gespeichert: {path}")))
            page.update()
        plt.close(fig)

    async def _save_eigentumsform_png(_e):
        fig = _create_eigentumsform_chart(state["df_gebaeude"])
        path = await ft.FilePicker().save_file(
            file_name="eigentumsform.png",
            allowed_extensions=["png"],
            file_type=ft.FilePickerFileType.CUSTOM,
        )
        if path:
            fig.savefig(path, format="png", dpi=150, bbox_inches="tight")
            page.show_dialog(ft.SnackBar(content=ft.Text(f"Gespeichert: {path}")))
            page.update()
        plt.close(fig)

    async def _save_surplus_deficit_pngs(_e):
        _, _, df_sd = get_results()
        figs = _create_surplus_deficit_charts(df_sd)

        path = await ft.FilePicker().save_file(
            file_name=f"ueberschuss_defizit_{state['scenario']}.zip",
            allowed_extensions=["zip"],
            file_type=ft.FilePickerFileType.CUSTOM,
        )
        if path:
            zip_bytes = _build_png_zip(
                [
                    (f"ueberschuss_defizit_{year}_{state['scenario']}.png", fig)
                    for year, fig in figs
                ]
            )
            with open(path, "wb") as f:
                f.write(zip_bytes)
            page.show_dialog(ft.SnackBar(content=ft.Text(f"Gespeichert: {path}")))
            page.update()

        for _, fig in figs:
            plt.close(fig)

    async def _save_all_pngs_zip(_e):
        _, df_demand, df_sd = get_results()
        fig_students = _create_students_chart(state["df_studierende"])
        fig_demand = _create_demand_chart(df_demand, state["scenario"])
        fig_eigentumsform = _create_eigentumsform_chart(state["df_gebaeude"])
        sd_figs = _create_surplus_deficit_charts(df_sd)

        path = await ft.FilePicker().save_file(
            file_name=f"diagramme_{state['scenario']}.zip",
            allowed_extensions=["zip"],
            file_type=ft.FilePickerFileType.CUSTOM,
        )
        if path:
            zip_bytes = _build_png_zip(
                [
                    ("studierende.png", fig_students),
                    (f"flaechenbedarf_{state['scenario']}.png", fig_demand),
                    ("eigentumsform.png", fig_eigentumsform),
                    *[
                        (f"ueberschuss_defizit_{year}_{state['scenario']}.png", fig)
                        for year, fig in sd_figs
                    ],
                ]
            )
            with open(path, "wb") as f:
                f.write(zip_bytes)
            page.show_dialog(ft.SnackBar(content=ft.Text(f"Gespeichert: {path}")))
            page.update()

        plt.close(fig_students)
        plt.close(fig_demand)
        plt.close(fig_eigentumsform)
        for _, fig in sd_figs:
            plt.close(fig)

    # ── Scenario dropdown ─────────────────────────────────────────────────

    def _scenario_options() -> list[ft.dropdown.Option]:
        """Return dropdown options derived from the loaded factors DataFrame."""
        if state["df_faktoren"] is not None:
            return [
                ft.dropdown.Option(s)
                for s in sorted(state["df_faktoren"]["Szenario"].unique().tolist())
            ]
        return []

    def _on_scenario_changed(e: ft.ControlEvent):
        state["scenario"] = e.control.value
        rebuild_content()
        page.update()

    scenario_dropdown = ft.Dropdown(
        label="Szenario wählen",
        options=_scenario_options(),
        value=state["scenario"],
        on_select=_on_scenario_changed,
        width=220,
    )

    def _update_scenario_options():
        """Sync the scenario dropdown with the currently loaded factors data.

        If the previously selected scenario is no longer available after a
        file change, the first available scenario is selected automatically.
        """
        scenario_dropdown.options = _scenario_options()
        if state["df_faktoren"] is not None:
            available = state["df_faktoren"]["Szenario"].unique().tolist()
            if state["scenario"] not in available:
                state["scenario"] = available[0] if available else "Basis"
                scenario_dropdown.value = state["scenario"]

    # ── Content container ─────────────────────────────────────────────────

    content_area = ft.Container(expand=True, padding=0)

    def rebuild_content() -> None:
        """Rebuild all tab content from current state."""
        log.debug("Rebuilding content with scenario: %s", state["scenario"])
        log.debug("Factors data:\n%s", state["df_faktoren"])
        log.debug("Student data:\n%s", state["df_studierende"])
        log.debug("Building data:\n%s", state["df_gebaeude"])
        if any(
            state[k] is None for k in ("df_gebaeude", "df_studierende", "df_faktoren")
        ):
            log.debug("Missing data, showing placeholder")
            content_area.content = ft.Container(
                content=ft.Text(
                    "Bitte zuerst Daten laden, Szenario wählen und berechnen.",
                    size=16,
                    color=ft.Colors.BLUE,
                ),
                padding=40,
            )
            return

        log.debug("Calculating results...")
        _, df_demand, df_sd = get_results()
        df_sd_display = _apply_nutzungsart_labels(df_sd)

        df_studierende_display = state["df_studierende"].rename(
            columns={
                "Forschung_Monatslohn": "Forschung (Monatslohn)",
                "Forschung_Stundenlohn": "Forschung (Stundenlohn)",
                "Services_Monatslohn": "Services (Monatslohn)",
                "Services_Stundenlohn": "Services (Stundenlohn)",
            }
        )

        # ── Tab 1: Übersicht ──────────────────────────────────────────────
        tab1 = ft.Column(
            [
                ft.Text("Gebäude & Räume", size=20, weight=ft.FontWeight.BOLD),
                ft.Container(
                    content=ft.Column(
                        [_df_to_datatable(state["df_gebaeude"])],
                        scroll=ft.ScrollMode.AUTO,
                    ),
                    height=300,
                ),
                ft.Divider(),
                ft.Text(
                    "Studierende & Mitarbeitende",
                    size=20,
                    weight=ft.FontWeight.BOLD,
                ),
                ft.Container(
                    content=ft.Column(
                        [
                            _df_to_datatable(df_studierende_display),
                        ],
                        scroll=ft.ScrollMode.AUTO,
                    ),
                    height=300,
                ),
                ft.Divider(),
                ft.Text(
                    f"Nutzungsfaktoren – Szenario: {state['scenario']}",
                    size=20,
                    weight=ft.FontWeight.BOLD,
                ),
                ft.Container(
                    content=ft.Column(
                        [
                            _df_to_datatable(
                                _apply_nutzungsart_labels(
                                    state["df_faktoren"][
                                    state["df_faktoren"]["Szenario"]
                                    == state["scenario"]
                                    ]
                                ),
                                round_to=0.1,
                            )
                        ],
                        scroll=ft.ScrollMode.AUTO,
                    ),
                    height=300,
                ),
            ],
            scroll=ft.ScrollMode.AUTO,
            spacing=16,
        )

        # ── Tab 2: Ergebnisse ─────────────────────────────────────────────
        metric_cards = []
        for year in [2026, 2030, 2040, 2050]:
            total_diff = df_sd[df_sd["Jahr"] == year]["Differenz_m2"].sum()
            metric_cards.append(
                _metric_card(
                    label=str(year),
                    value=f"{total_diff:,.0f} m²",
                    is_surplus=total_diff >= 0,
                )
            )

        # Pivot table with colour coding
        pivot = df_sd_display.pivot_table(
            index="Nutzungsart",
            columns="Jahr",
            values="Differenz_m2",
            aggfunc="first",
        )
        pivot = pivot.reindex(columns=[2026, 2030, 2040, 2050], fill_value=0)
        pivot.columns = [str(c) for c in pivot.columns]

        pivot_columns = [
            ft.DataColumn(ft.Text("Nutzungsart", weight=ft.FontWeight.BOLD))
        ] + [
            ft.DataColumn(ft.Text(col, weight=ft.FontWeight.BOLD))
            for col in pivot.columns
        ]
        pivot_rows = []
        for nutzungsart, row in pivot.iterrows():
            cells: list[ft.DataCell] = [ft.DataCell(ft.Text(str(nutzungsart)))]
            for val in row:
                try:
                    v = float(val)
                    tc = "#276221" if v > 0 else ("#9c0006" if v < 0 else None)
                    cells.append(
                        ft.DataCell(
                            ft.Container(
                                content=ft.Text(_format_display_value(v), color=tc),
                                padding=8,
                            )
                        )
                    )
                except (TypeError, ValueError):
                    cells.append(ft.DataCell(ft.Text(str(val))))
            pivot_rows.append(ft.DataRow(cells=cells))

        pivot_table = ft.DataTable(
            columns=pivot_columns,
            rows=pivot_rows,
            border=ft.Border.all(1, ft.Colors.GREY_300),
            heading_row_color=ft.Colors.BLUE_GREY_50,
            horizontal_lines=ft.BorderSide(1, ft.Colors.GREY_200),
        )

        df_sd_display_formatted = df_sd_display.copy()
        df_sd_display_formatted["Fläche"] = (
            df_sd_display_formatted["Fläche"].map("{:,.0f}".format).str.replace(",", "'")
        )
        df_sd_display_formatted["Bedarf_m2"] = (
            df_sd_display_formatted["Bedarf_m2"].map("{:,.0f}".format).str.replace(",", "'")
        )
        df_sd_display_formatted["Differenz_m2"] = (
            df_sd_display_formatted["Differenz_m2"].map("{:,.0f}".format).str.replace(",", "'")
        )
        df_sd_display_formatted = df_sd_display_formatted.rename(
            columns={
                "Fläche": "Ist-Fläche (m²)",
                "Bedarf_m2": "Bedarf (m²)",
                "Differenz_m2": "Differenz (m²)",
            }
        )

        tab2 = ft.Column(
            [
                ft.Text(
                    f"Flächen-Überschuss/Defizit – Szenario: {state['scenario']}",
                    size=20,
                    weight=ft.FontWeight.BOLD,
                ),
                ft.Row(metric_cards, spacing=16, wrap=False, scroll=ft.ScrollMode.AUTO),
                ft.Divider(),
                ft.Text(
                    "Differenz (m²) pro Nutzungsart und Jahr "
                    "(grün = Überschuss, rot = Defizit)",
                    weight=ft.FontWeight.BOLD,
                ),
                ft.Container(
                    content=ft.Column([pivot_table], scroll=ft.ScrollMode.AUTO),
                ),
                ft.Divider(),
                ft.Text(
                    "Vollständige Ergebnistabelle",
                    size=20,
                    weight=ft.FontWeight.BOLD,
                ),
                ft.Container(
                    content=ft.Column(
                        [_df_to_datatable(df_sd_display_formatted)],
                        scroll=ft.ScrollMode.AUTO,
                    ),
                    height=300,
                ),
            ],
            scroll=ft.ScrollMode.AUTO,
            spacing=16,
        )

        # ── Tab 3: Diagramme ─────────────────────────────────────────────
        fig_students = _create_students_chart(state["df_studierende"])
        fig_demand = _create_demand_chart(df_demand, state["scenario"])
        fig_eigentumsform = _create_eigentumsform_chart(state["df_gebaeude"])
        sd_figs = _create_surplus_deficit_charts(df_sd)

        sd_chart_controls: list[ft.Control] = [
            ft.Container(
                content=ft.Image(src=_fig_to_base64(fig)),
                expand=True,
            )
            for _, fig in sd_figs
        ]

        num_cols = 2
        sd_chart_rows: list[ft.Control] = [
            ft.Row(
                controls=sd_chart_controls[i : i + num_cols],
                spacing=16,
            )
            for i in range(0, len(sd_chart_controls), num_cols)
        ]
        sd_charts_column = ft.Column(controls=sd_chart_rows, spacing=16)

        tab3 = ft.Column(
            controls=[
                ft.Text(
                    "Studierendenzahlen & Kategorien im Zeitverlauf",
                    size=20,
                    weight=ft.FontWeight.BOLD,
                ),
                ft.Container(
                    content=ft.Image(src=_fig_to_base64(fig_students)),
                    alignment=ft.Alignment.CENTER,
                ),
                ft.Divider(),
                ft.Text(
                    "Flächenbedarf nach Nutzungsart und Jahr",
                    size=20,
                    weight=ft.FontWeight.BOLD,
                ),
                ft.Container(
                    content=ft.Image(src=_fig_to_base64(fig_demand)),
                    alignment=ft.Alignment.CENTER,
                ),
                ft.Divider(),
                ft.Text(
                    "Fläche nach Eigentumsform",
                    size=20,
                    weight=ft.FontWeight.BOLD,
                ),
                ft.Container(
                    content=ft.Image(src=_fig_to_base64(fig_eigentumsform)),
                    alignment=ft.Alignment.CENTER,
                ),
                ft.Divider(),
                ft.Text(
                    "Überschuss/Defizit nach Nutzungsart",
                    size=20,
                    weight=ft.FontWeight.BOLD,
                ),
                sd_charts_column,
            ],
            scroll=ft.ScrollMode.AUTO,
            spacing=16,
        )

        # Close all matplotlib figures to free memory
        plt.close("all")

        # ── Tab 4: Export ─────────────────────────────────────────────────
        tab4 = ft.Column(
            [
                ft.Text(
                    "Ergebnisse exportieren",
                    size=20,
                    weight=ft.FontWeight.BOLD,
                ),
                ft.Button(
                    "📥 Ergebnisse als Excel speichern",
                    on_click=_save_excel,
                    icon=ft.Icons.SAVE,
                ),
                ft.Button(
                    "📥 Gerundete Flächen als Excel speichern",
                    on_click=_save_excel_rounded,
                    icon=ft.Icons.SAVE,
                ),
                ft.Divider(),
                ft.Text(
                    "Diagramme als PNG speichern",
                    weight=ft.FontWeight.BOLD,
                ),
                ft.Row(
                    [
                        ft.Button(
                            "📥 Studierendenzahlen (PNG)",
                            on_click=_save_students_png,
                            icon=ft.Icons.IMAGE,
                        ),
                        ft.Button(
                            "📥 Flächenbedarf (PNG)",
                            on_click=_save_demand_png,
                            icon=ft.Icons.IMAGE,
                        ),
                        ft.Button(
                            "📥 Eigentumsform (PNG)",
                            on_click=_save_eigentumsform_png,
                            icon=ft.Icons.IMAGE,
                        ),
                        ft.Button(
                            "📥 Überschuss/Defizit (ZIP)",
                            on_click=_save_surplus_deficit_pngs,
                            icon=ft.Icons.FOLDER_ZIP,
                        ),
                    ],
                    spacing=16,
                ),
                ft.Button(
                    "📦 Alle Diagramme (ZIP)",
                    on_click=_save_all_pngs_zip,
                    icon=ft.Icons.FOLDER_ZIP,
                ),
            ],
            spacing=16,
        )

        # ── Assemble tabs ────────────────────────────────────────────────

        content_area.content = ft.Tabs(
            selected_index=0,
            length=4,
            expand=True,
            content=ft.Column(
                expand=True,
                controls=[
                    ft.TabBar(
                        tabs=[
                            ft.Tab(label=ft.Text("📋 Übersicht")),
                            ft.Tab(label=ft.Text("📊 Ergebnisse")),
                            ft.Tab(label=ft.Text("📈 Diagramme")),
                            ft.Tab(label=ft.Text("⬇️ Export")),
                        ]
                    ),
                    ft.TabBarView(
                        expand=True,
                        controls=[
                            ft.Container(content=tab1, padding=20),
                            ft.Container(content=tab2, padding=20),
                            ft.Container(content=tab3, padding=20),
                            ft.Container(content=tab4, padding=20),
                        ],
                    ),
                ],
            ),
        )

    # ── Sidebar ───────────────────────────────────────────────────────────

    sidebar = ft.Container(
        content=ft.Column(
            [
                ft.Text(
                    "⚙️ Einstellungen",
                    size=18,
                    weight=ft.FontWeight.BOLD,
                ),
                ft.Divider(),
                ft.Text(
                    "📂 Eigene Dateien laden",
                    weight=ft.FontWeight.BOLD,
                    size=14,
                ),
                ft.Button(
                    "Gebäude & Räume",
                    on_click=load_inventory_file,
                    icon=ft.Icons.UPLOAD_FILE,
                    width=220,
                ),
                gebaeude_label,
                ft.Button(
                    "Studierende & Mitarbeitende",
                    on_click=load_person_file,
                    icon=ft.Icons.UPLOAD_FILE,
                    width=220,
                ),
                studierende_label,
                ft.Button(
                    "Nutzungsfaktoren",
                    on_click=load_faktoren_file,
                    icon=ft.Icons.UPLOAD_FILE,
                    width=220,
                ),
                faktoren_label,
                ft.Container(
                    content=ft.Button(
                        "Daten laden",
                        on_click=load_all_data,
                        icon=ft.Icons.CALCULATE,
                        width=220,
                    ),
                    padding=ft.Padding(0, 20),
                ),
                ft.Divider(),
                ft.Text("Szenario", weight=ft.FontWeight.BOLD),
                scenario_dropdown,
                ft.Button(
                    "Szenario berechnen",
                    on_click=run_calculations,
                    icon=ft.Icons.CALCULATE,
                    width=220,
                ),
            ],
            spacing=8,
            scroll=ft.ScrollMode.AUTO,
        ),
        width=280,
        padding=20,
        bgcolor=ft.Colors.GREY_50,
        border=ft.Border.only(right=ft.BorderSide(1, ft.Colors.GREY_300)),
    )

    # ── Initial data load & render ────────────────────────────────────────

    _update_scenario_options()
    rebuild_content()

    page.add(
        ft.Row(
            [sidebar, content_area],
            expand=True,
            vertical_alignment=ft.CrossAxisAlignment.START,
        )
    )


if __name__ == "__main__":
    ft.run(main, assets_dir="assets")
