"""Raumprognose Tool – Streamlit Dashboard.

Run with:
    streamlit run app/streamlit_app.py
"""

from __future__ import annotations

import io
from typing import Any

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from calculations import current_area_by_nutzungsart, future_demand, surplus_deficit
from data_loader import load_gebaeude_raeume, load_nutzungsfaktoren, load_studierende

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(page_title="Raumprognose Tool", layout="wide")
st.title("🏛️ Raumprognose Tool")

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Einstellungen")

    st.subheader("Szenario")
    # Load default factors to populate the scenario dropdown
    try:
        _df_nf_default = load_nutzungsfaktoren()
        available_scenarios = sorted(_df_nf_default["szenario"].unique().tolist())
    except Exception:
        available_scenarios = ["Basis"]

    selected_scenario = st.selectbox(
        "Szenario wählen",
        options=available_scenarios,
        index=0,
    )

    st.divider()
    st.subheader("📂 Eigene Dateien hochladen (optional)")
    upload_gebaeude = st.file_uploader(
        "Gebäude & Räume (.xlsx)", type=["xlsx"], key="upload_gebaeude"
    )
    upload_studierende = st.file_uploader(
        "Studierende (.xlsx)", type=["xlsx"], key="upload_studierende"
    )
    upload_faktoren = st.file_uploader(
        "Nutzungsfaktoren (.xlsx)", type=["xlsx"], key="upload_faktoren"
    )


# ── Load data ─────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def _load_all(
    gb_bytes: bytes | None,
    st_bytes: bytes | None,
    nf_bytes: bytes | None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Load all three DataFrames, preferring uploaded files over defaults."""
    df_gb = load_gebaeude_raeume(io.BytesIO(gb_bytes) if gb_bytes else None)
    df_st = load_studierende(io.BytesIO(st_bytes) if st_bytes else None)
    df_nf = load_nutzungsfaktoren(io.BytesIO(nf_bytes) if nf_bytes else None)
    return df_gb, df_st, df_nf


try:
    df_gebaeude, df_studierende, df_faktoren = _load_all(
        upload_gebaeude.read() if upload_gebaeude else None,
        upload_studierende.read() if upload_studierende else None,
        upload_faktoren.read() if upload_faktoren else None,
    )
except Exception as exc:
    st.error(
        f"Fehler beim Laden der Daten: {exc}\n\n"
        "Bitte stellen Sie sicher, dass die Excel-Dateien unter `data/` vorhanden sind "
        "oder laden Sie eigene Dateien in der Sidebar hoch."
    )
    st.stop()

# ── Calculations ──────────────────────────────────────────────────────────────
df_current = current_area_by_nutzungsart(df_gebaeude)
df_demand = future_demand(df_studierende, df_faktoren, selected_scenario)
df_sd = surplus_deficit(df_current, df_demand)

# ── Tabs ──────────────────────────────────────────────────────────────────────
tab_overview, tab_results, tab_charts, tab_export = st.tabs(
    ["📋 Übersicht", "📊 Ergebnisse", "📈 Diagramme", "⬇️ Export"]
)

# ════════════════════════════════════════════════════════════════════════════════
# Tab 1 – Overview / Input data
# ════════════════════════════════════════════════════════════════════════════════
with tab_overview:
    st.subheader("Gebäude & Räume")
    st.dataframe(df_gebaeude, use_container_width=True)

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Studierendenzahlen")
        st.dataframe(df_studierende, use_container_width=True)
    with col2:
        st.subheader(f"Nutzungsfaktoren – Szenario: {selected_scenario}")
        st.dataframe(
            df_faktoren[df_faktoren["szenario"] == selected_scenario],
            use_container_width=True,
        )

# ════════════════════════════════════════════════════════════════════════════════
# Tab 2 – Results
# ════════════════════════════════════════════════════════════════════════════════
with tab_results:
    st.subheader(f"Flächen-Über-/Unterschuss – Szenario: {selected_scenario}")

    # Summary metrics
    years = sorted(df_sd["jahr"].unique())
    metric_cols = st.columns(len(years))
    for col, year in zip(metric_cols, years):
        total_diff = df_sd[df_sd["jahr"] == year]["differenz_m2"].sum()
        col.metric(
            label=str(year),
            value=f"{total_diff:,.0f} m²",
            delta=f"{'Überschuss' if total_diff >= 0 else 'Defizit'}",
        )

    st.divider()

    # Detailed table with colour coding
    st.write(
        "**Differenz (m²) pro Nutzungsart und Jahr** (grün = Überschuss, rot = Defizit)"
    )

    def _color_diff(val: Any) -> str:
        """Return a CSS background-color string based on the sign of *val*."""
        try:
            v = float(val)
        except (TypeError, ValueError):
            return ""
        if v > 0:
            return "background-color: #c6efce; color: #276221"
        if v < 0:
            return "background-color: #ffc7ce; color: #9c0006"
        return ""

    pivot = df_sd.pivot_table(
        index="nutzungsart",
        columns="jahr",
        values="differenz_m2",
        aggfunc="first",
    )
    pivot.columns = [str(c) for c in pivot.columns]
    styled = pivot.style.map(_color_diff).format("{:.1f}")
    st.dataframe(styled, use_container_width=True)

    st.divider()
    st.subheader("Vollständige Ergebnistabelle")
    st.dataframe(
        df_sd.rename(
            columns={
                "nutzungsart": "Nutzungsart",
                "jahr": "Jahr",
                "flaeche_m2_gesamt": "Ist-Fläche (m²)",
                "bedarf_m2": "Bedarf (m²)",
                "differenz_m2": "Differenz (m²)",
            }
        ),
        use_container_width=True,
    )

# ════════════════════════════════════════════════════════════════════════════════
# Tab 3 – Charts
# ════════════════════════════════════════════════════════════════════════════════
with tab_charts:
    # 1. Line chart – student numbers over time
    st.subheader("Studierendenzahlen im Zeitverlauf")
    fig_students = px.line(
        df_studierende,
        x="jahr",
        y="anzahl_studierende",
        markers=True,
        labels={"jahr": "Jahr", "anzahl_studierende": "Studierende"},
        title="Entwicklung der Studierendenzahlen",
    )
    st.plotly_chart(fig_students, use_container_width=True)

    # 2. Grouped bar chart – area demand by usage type and year
    st.subheader("Flächenbedarf nach Nutzungsart und Jahr")
    fig_demand = px.bar(
        df_demand,
        x="nutzungsart",
        y="bedarf_m2",
        color="jahr",
        barmode="group",
        labels={
            "nutzungsart": "Nutzungsart",
            "bedarf_m2": "Bedarf (m²)",
            "jahr": "Jahr",
        },
        title=f"Flächenbedarf – Szenario: {selected_scenario}",
        color_continuous_scale="Blues",
    )
    st.plotly_chart(fig_demand, use_container_width=True)

    # 3. Bar chart – surplus / deficit by usage type
    st.subheader("Über-/Unterschuss nach Nutzungsart")
    fig_cols = st.columns(len(years))
    for col, year in zip(fig_cols, years):
        df_year = df_sd[df_sd["jahr"] == year].copy()
        df_year["Farbe"] = df_year["differenz_m2"].apply(
            lambda x: "Überschuss" if x >= 0 else "Defizit"
        )
        fig_diff = px.bar(
            df_year,
            x="nutzungsart",
            y="differenz_m2",
            color="Farbe",
            color_discrete_map={"Überschuss": "#2ecc71", "Defizit": "#e74c3c"},
            labels={"nutzungsart": "Nutzungsart", "differenz_m2": "Differenz (m²)"},
            title=str(year),
        )
        fig_diff.update_layout(showlegend=False)
        col.plotly_chart(fig_diff, use_container_width=True)

# ════════════════════════════════════════════════════════════════════════════════
# Tab 4 – Export
# ════════════════════════════════════════════════════════════════════════════════
with tab_export:
    st.subheader("Ergebnisse exportieren")

    # ── Excel export ──────────────────────────────────────────────────────────
    def _build_excel(
        df_results: pd.DataFrame,
        df_stud: pd.DataFrame,
        df_dem: pd.DataFrame,
    ) -> bytes:
        """Build a styled multi-sheet Excel workbook and return it as bytes.

        Args:
            df_results: Full surplus/deficit DataFrame.
            df_stud: Student numbers DataFrame.
            df_dem: Future demand DataFrame.

        Returns:
            Raw bytes of the ``.xlsx`` workbook.
        """
        wb = Workbook()

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F4E79")
        green_fill = PatternFill("solid", fgColor="C6EFCE")
        red_fill = PatternFill("solid", fgColor="FFC7CE")
        center = Alignment(horizontal="center")

        def _write_sheet(
            ws: Any, df: pd.DataFrame, title: str, diff_col: str | None = None
        ) -> None:
            ws.title = title
            for r_idx, row in enumerate(
                dataframe_to_rows(df, index=False, header=True), start=1
            ):
                ws.append(row)
                if r_idx == 1:
                    for cell in ws[r_idx]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center
                elif diff_col and r_idx > 1:
                    col_names = list(df.columns)
                    if diff_col in col_names:
                        c_idx = col_names.index(diff_col) + 1
                        val = ws.cell(row=r_idx, column=c_idx).value
                        try:
                            if float(val) >= 0:
                                ws.cell(row=r_idx, column=c_idx).fill = green_fill
                            else:
                                ws.cell(row=r_idx, column=c_idx).fill = red_fill
                        except (TypeError, ValueError):
                            pass

        ws1 = wb.active
        _write_sheet(ws1, df_results, "Ergebnisse", diff_col="differenz_m2")

        ws2 = wb.create_sheet()
        _write_sheet(ws2, df_stud, "Studierende")

        ws3 = wb.create_sheet()
        _write_sheet(ws3, df_dem, "Flächenbedarf")

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    excel_bytes = _build_excel(df_sd, df_studierende, df_demand)
    st.download_button(
        label="📥 Ergebnisse als Excel herunterladen",
        data=excel_bytes,
        file_name=f"raumprognose_{selected_scenario}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.divider()

    # ── Chart PNG export ──────────────────────────────────────────────────────
    st.write("**Diagramme als PNG herunterladen**")

    dl_cols = st.columns(2)
    with dl_cols[0]:
        st.download_button(
            label="📥 Studierendenzahlen (PNG)",
            data=fig_students.to_image(format="png"),
            file_name="studierende.png",
            mime="image/png",
        )
    with dl_cols[1]:
        st.download_button(
            label="📥 Flächenbedarf (PNG)",
            data=fig_demand.to_image(format="png"),
            file_name=f"flaechenbedarf_{selected_scenario}.png",
            mime="image/png",
        )
