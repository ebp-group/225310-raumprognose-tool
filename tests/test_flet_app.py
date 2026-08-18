import io

import pandas as pd
from openpyxl import load_workbook

from app.flet_app import _build_excel_rounded


def _build_test_workbook():
    df_sd = pd.DataFrame(
        {
            "Nutzungsart": ["Sport", "Sport", "Büro", "Büro"],
            "Jahr": [2030, 2040, 2030, 2040],
            "Fläche": [49.0, 81.0, 102.0, 153.0],
            "Bedarf_m2": [76.0, 125.0, 151.0, 195.0],
            "Differenz_m2": [-27.0, -44.0, -49.0, -42.0],
        }
    )
    df_studierende = pd.DataFrame(
        {
            "Jahr": [2030, 2030, 2040, 2040],
            "Kategorie": ["Studierende", "Forschung", "Studierende", "Forschung"],
            "Anzahl": [500, 50, 600, 60],
        }
    )
    workbook = load_workbook(io.BytesIO(_build_excel_rounded(df_sd, df_studierende)))
    return workbook["Gerundete Flächen"]


def test_build_excel_rounded_header_row() -> None:
    sheet = _build_test_workbook()

    assert sheet.cell(row=1, column=1).value == "Raumtypen"
    assert sheet.cell(row=1, column=2).value == 2030
    assert sheet.cell(row=1, column=2).alignment.horizontal == "left"


def test_build_excel_rounded_appends_subtotal_and_headcount_rows() -> None:
    sheet = _build_test_workbook()

    labels = [sheet.cell(row=r, column=1).value for r in range(1, sheet.max_row + 1)]
    assert "Total Lehre HNF 1 / 3 / 5" in labels
    assert "Anzahl Studierende" in labels
    assert "Total Büro HNF 2" in labels
    assert "Total Lehre und Büro HNF 1 / 2 / 3 / 5" in labels
    # "Anzahl Mitarbeitende" comes right after "Total Büro HNF 2"; the grand
    # total is the very last row.
    total_buero_idx = labels.index("Total Büro HNF 2")
    assert labels[total_buero_idx + 1] == "Anzahl Mitarbeitende"
    assert labels[-1] == "Total Lehre und Büro HNF 1 / 2 / 3 / 5"

    row_by_label = {label: r for r, label in enumerate(labels, start=1) if label}

    # Grand total: Sport (IST=50, SOLL=75, Diff=-25) + Büro (IST=100, SOLL=150, Diff=-50) for 2030.
    grand_total_row = row_by_label["Total Lehre und Büro HNF 1 / 2 / 3 / 5"]
    assert sheet.cell(row=grand_total_row, column=2).value == 150
    assert sheet.cell(row=grand_total_row, column=3).value == 225
    assert sheet.cell(row=grand_total_row, column=4).value == -75
    assert sheet.cell(row=grand_total_row, column=5).value == 235
    assert sheet.cell(row=grand_total_row, column=6).value == 320
    assert sheet.cell(row=grand_total_row, column=7).value == -85

    # Total Lehre only contains Sport in this fixture.
    total_lehre_row = row_by_label["Total Lehre HNF 1 / 3 / 5"]
    assert sheet.cell(row=total_lehre_row, column=2).value == 50
    assert sheet.cell(row=total_lehre_row, column=3).value == 75
    assert sheet.cell(row=total_lehre_row, column=4).value == -25

    # Total Büro mirrors the single Büro row.
    total_buero_row = row_by_label["Total Büro HNF 2"]
    assert sheet.cell(row=total_buero_row, column=2).value == 100
    assert sheet.cell(row=total_buero_row, column=3).value == 150
    assert sheet.cell(row=total_buero_row, column=4).value == -50

    # Headcounts land in the SOLL column only.
    studierende_row = row_by_label["Anzahl Studierende"]
    assert sheet.cell(row=studierende_row, column=2).value is None
    assert sheet.cell(row=studierende_row, column=3).value == 500
    assert sheet.cell(row=studierende_row, column=4).value is None
    assert sheet.cell(row=studierende_row, column=1).font.italic is True

    mitarbeitende_row = row_by_label["Anzahl Mitarbeitende"]
    assert sheet.cell(row=mitarbeitende_row, column=3).value == 50
